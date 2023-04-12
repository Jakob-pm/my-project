# date = 03/29/2023
# author = Jakob
print('hello world!')


# 11 生成excel工资条

# '/Users/jakob_pc/Desktop/Python_datas/103_day7/data11/salary_list.xlsx'

# from openpyxl import load_workbook, Workbook
#
# def creat_excel():
#     wb = load_workbook(ori_file, data_only = True)
#     sh1 = wb.active
#     # title = sh1.rows
#     # lst1 = list(sh1.rows)
#     # lst2 =list(list(sh1.rows)[0])
#     title = [cell.value for cell in list(list(sh1.rows)[0])]
#     for i, row in enumerate(sh1.rows):
#         if i == 0:
#             continue
#         else:
#             temp_wb = Workbook()
#             temp_sh = temp_wb.active
#             row = [cell.value for cell in row]
#             print(row)
#             temp_sh.append(title)
#             temp_sh.append(row)
#             temp_wb.save(f'{file_path}/{row[1]}.xlsx')
#
#
#
# file_path = '/Users/jakob_pc/Desktop/Python_datas/103_day7/data11'
# ori_file ='/Users/jakob_pc/Desktop/Python_datas/103_day7/data11/salary_list.xlsx'
# creat_excel()
#
# print('done!')

# 12 隔行变色

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
# from xlutils.copy import copy
# import xlrd

def creat_excel():
    wb = load_workbook(ori_file, data_only = True)
    sh1 = wb.active

    back_color = PatternFill('solid',fgColor = 'AEEEEE')
    n = 0
    for i in range(2, sh1.max_row+1,2):
        #if i %2 == 0:
        for cell in range(1, sh1.max_column+1):
            sh1.cell(i,cell).fill = back_color
        print('*')
        n +=1
    wb.save(f'{file_path}/output.xlsx')
    print(n)
    print('done!')
ori_file ='/Users/jakob_pc/Desktop/Python_datas/103_day7/data12.xlsx'
file_path = '/Users/jakob_pc/Desktop/Python_datas/103_day7'
creat_excel()




# if __name__ == '__main__':
    # main()

print('Goodbye')