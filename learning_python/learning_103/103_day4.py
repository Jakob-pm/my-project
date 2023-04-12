# date = 03122023
# author = Jakob
print('hello world!')


# 08 openpyxl创建excel

def new_sheet():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh2 = wb.create_sheet()
    sh3 = wb.create_sheet('数据')
    sh4 = wb.create_sheet('人员',0)
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')
    print(sh1) # <Worksheet "Sheet">
    print(sh2) # <Worksheet "Sheet1">
    print(sh3)
    print(sh4)

def set_value():
    from openpyxl.styles import Font, Alignment, colors
    bold_italic_30_font = Font(name = '微软雅黑', size = 30, bold = True, color = colors.BLUE)
    bold_italic_20_font = Font(name='微软雅黑', size=20, italic=True, underline='singleAccounting', bold=True, color='CD1076')
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1['A1'] = 'Hello world!'
    sh1['A1'].font = bold_italic_30_font
    sh1['B1'] = 'Python!'
    sh1['B1'].font = bold_italic_20_font
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')


# new_sheet()
set_value()



# 09 多个Excel合并为一个excel

# 10 多个Excel合并为一个excel

# 11 生成excel工资条




# if __name__ == '__main__':
    # main()

print('Goodbye')



