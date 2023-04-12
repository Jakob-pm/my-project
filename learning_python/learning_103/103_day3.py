# date = 03122023
# author = Jakob
print('hello world!')

# 06 表格的拆分

'''import xlrd
from xlutils.copy import copy

def get_data():
    wb = xlrd.open_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分.xls')
    sh = wb.sheet_by_index(0)
    # print(sh)
    all_data = {}
    for r in range(sh.nrows):
        d = {'type':sh.cell_value(r,1),'item':sh.cell_value(r,2),'amount':sh.cell_value(r,3),'sum':sh.cell_value(r,4)}
        # print(d)
        key = sh.cell_value(r,0)
        if all_data.get(key):
            all_data[key].append(d)
        else:
            all_data[key]=[d]
    # print(all_data)
    return all_data



def save(data):
    wb = xlrd.open_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分.xls')
    wb2 = copy(wb)
    for key in data.keys():
        temp = wb2.add_sheet(key)
        for i,d in enumerate(data.get(key)):
            temp.write(i,0,d.get('type'))
            temp.write(i, 1, d.get('item'))
            temp.write(i, 2, d.get('amount'))
            temp.write(i, 3, d.get('sum'))
    wb2.save('/Users/jakob_pc/Desktop/Python实践/表格拆分02.xls')


data = get_data()
save(data)
'''

# 07 openpyxl读取数据

'''def open():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    # print(wb)
    # 获取工作表的3种方法
    sh1 = wb.active
    sh2 = wb['A']
    sh3 = wb.get_sheet_by_name('A')
    # print(sh1)
    # print(sh2)
    # print(sh3)
# open()

def show_all_sheets():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)

# show_all_sheets()

def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    sh1 = wb.active
    value1 = sh1.cell(8,3)
    value2 = sh1['C8']
    print(value1)
    print(value2)
    value1 = sh1.cell(8, 3).value
    value2 = sh1['C8'].value
    print(value1)
    print(value2)
# get_one_value()

def get_multi_values():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    sh1 = wb['Sheet1']
    # 切片
    cells1 = sh1['b2':'e11']
    print(type(cells1)) # 元组类型，元组里每行的值组成一个元组
    data = {}
    key = 0
    for i in cells1:
        key += 1 # 有没有什么办法可以让key显示行数
        d = {}
        lst1 =[]
        for k in i:
            lst1.append(k.value)
        d['type']=lst1[0]
        d['item'] = lst1[1]
        d['amount'] = lst1[2]
        d['sum'] = lst1[3]
        data[key]= d # 转化为字典更方便查询，转化为list更容易看
    # print(data)
    # 整行、整列
    cell_row3 = sh1[3]
    cell_col2 = sh1['b']
    # print(cell_col2)
    print(type(cell_col2)) # 还是元组类型
    # print(cell_row3)
    # 整行整列数据也支持切片
    cell_row3_5 = sh1[3:5]
    # print(cell_row3_5)
    # 通过迭代获取数据
    for row in sh1.iter_rows(min_row =2, max_row=5,max_col=3):
        for cell in row:
            print(cell.value)


# get_multi_values()

# 获取表格内所有数据
def get_all_data():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    sh1 = wb['Sheet1']
    for row in sh1.rows:
        for cell in row:
            print(cell.value)

# get_all_data()

# 获取表格行数和列数
def get_num():
    from openpyxl import load_workbook
    wb = load_workbook('/Users/jakob_pc/Desktop/Python实践/表格拆分03.xlsx')
    sh1 = wb['Sheet1']
    print(sh1.max_row)
    print(sh1.max_column)
get_num()
'''
# 08 openpyxl创建excel

def new_sheet():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh2 = wb.create_sheet()
    print(sh1) # <Worksheet "Sheet">
    print(sh2) # <Worksheet "Sheet1">


# new_sheet()

# 09 多个Excel合并为一个excel

# 10 多个Excel合并为一个excel

# 11 生成excel工资条




# if __name__ == '__main__':
    # main()

print('Goodbye')