# 解决openpyxl里的copy问题

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
# from xlutils.copy import copy
# import xlrd

def creat_excel():
    wb = load_workbook(ori_file)
    sh1 = wb.active
    temp_sh = wb.copy_worksheet(sh1)
    print(type(temp_sh))
    back_color = PatternFill('solid',fgColor = 'AEEEEE')
    for i in range(1, temp_sh.max_row+1):
        if i %2 == 0:
            for cell in range(1, temp_sh.max_column+1):
                temp_sh.cell(i,cell).fill = back_color
                print('*')
    del wb['Sheet1']
    wb.save(f'{file_path}/output.xlsx')
    print('done!')



def creat_excel2():
    wb = load_workbook(ori_file)
    sh1 = wb.active

    back_color = PatternFill('solid',fgColor = 'AEEEEE')
    n = 0
    for i in range(2, sh1.max_row+1,2):
        #if i %2 == 0:
        for cell in range(1, sh1.max_column+1):
            sh1.cell(i,cell).fill = back_color
        print('*')
        n +=1
    wb.save(f'{file_path}/output.xlsx')
    print(n)
    print('done!')
ori_file ='/Users/jakob_pc/Desktop/Python_datas/103_day7/data12.xlsx'
file_path = '/Users/jakob_pc/Desktop/Python_datas/103_day7'
creat_excel2()


print('Goodbye')