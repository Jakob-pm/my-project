import openpyxl

# Load the existing workbook (replace "example.xlsx" with the name of your file)
workbook = openpyxl.load_workbook("/Users/jakob_pc/Desktop/Python实践/test_xl_file_02.xlsx")

# Get the active sheet (replace "Sheet1" with the name of your sheet)
sheet = workbook["A Test Sheet_01"]

# Calculate the row number of the last line in the sheet
last_row_num = sheet.max_row

# Add a new row with the text "hello world"
new_row = last_row_num + 1
sheet.cell(row=new_row, column=1).value = "hello world"

# Save the changes to the workbook
workbook.save("/Users/jakob_pc/Desktop/Python实践/test_xl_file_02.xlsx")