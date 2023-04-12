# date = 03/30/2023
# author = Jakob
print('hello world!')

# 13 快速统计加班时间

from openpyxl import Workbook, load_workbook
from datetime import date

def create_data():
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date','姓名','打卡时间'],
        [date(2023, 3, 12), '张飞', '18:59'],
        [date(2023, 3, 13), '关羽', '19:40'],
        [date(2023, 3, 14), '黄忠', '21:31'],
        [date(2023, 3, 15), '赵云', '17:30'],
        [date(2023, 3, 16), '马超', '15:50'],
        [date(2023, 3, 17), '诸葛亮', '22:47']
    ]
    for row in rows:
        sh.append(row)
    wb.save(ori_file)
    print('done!')

def overtime():
    wb = load_workbook(ori_file)
    sh1 = wb.active
    title = [cell.value for cell in list(list(sh1.rows)[0])]
    data = []
    for i in range(2, sh1.max_row+1):
        t_data = []
        for j in range(1,sh1.max_column+1):
            t_data.append(sh1.cell(i,j).value)
        h,m = t_data[2].split(':')
        full = int(h)*60 +int(m)
        temp = full - 18*60
        temp = 0 if temp < 0 else temp
        # print(temp)
        # print(type(temp))
        t_data.append(temp)
        t_data[0] = t_data[0].date() #日期格式的处理
        data.append(t_data)
    # print(data)

    n_wb = Workbook()
    n_sh1 = n_wb.active
    title.append('加班时长')
    n_sh1.append(title)
    for d in data:
        n_sh1.append(d)
        print('*')
    n_wb.save(tar_file)
    print('done!')



ori_file = '/Users/jakob_pc/Desktop/Python_datas/103_day8/data13.xlsx'
tar_file = '/Users/jakob_pc/Desktop/Python_datas/103_day8/data13_output.xlsx'

# create_data()
overtime()

# if __name__ == '__main__':
    # main()

print('Goodbye')