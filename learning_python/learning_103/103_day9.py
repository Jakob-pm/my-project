# date = 03/31/2023
# author = Jakob
print('hello world!')

# 14 快速查excel重复数据

# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
#
# def dum():
#     wb = load_workbook(file)
#     sh1 = wb.active
#
#     index = [] # 用来接收重复数据的索引
#     temp = []  # 用来接收非重复数据的内容
#     for i,c in enumerate(sh1['B']):
#         print(i, end= '')
#         c = c.value
#         if c not in temp:
#             temp.append(c)
#             print('*')
#         else:
#             index.append(i) #注意python中index从0开始，而excel中从1开始
#             print('$')
#     print(temp)
#     print(index)
#
#     fill = PatternFill('solid', fgColor ='AEEEEE')
#     for i,row in enumerate(sh1.rows): # enumerate从0开始
#         if i in index:
#             for cell in row:
#                 cell.fill = fill
#             print(f'第{i+1}条是重复数据')
#
#         wb.save(file)
#
# file = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data14.xlsx'
# dum()
# print('done!')

# 15 身份证号提前信息

from openpyxl import load_workbook
from datetime import datetime

# 第一种方法：写入单列数据
# def create_time():
#     wb= load_workbook(file)
#     sh1 = wb.active
#     lst1 = [cell.value for cell in sh1['D']]
#     lst1 = lst1[1:]
#     curr_year = datetime.now().year
#     print(curr_year)
#     print(type(curr_year))
#     # print(lst1)
#     for i in lst1:
#         year = int(i[6:10])
#         #print(type(year))
#         month =int(i[10:12])
#         date = int(i[12:14])
#         print(f'{year}年{month}月{date}日')
#         age = curr_year-year
#         # print(f'{age}age')
#         data.append(age)
#     return data
#
# def create_sheet():
#     wb = load_workbook(file,data_only = True)
#     sh1 = wb.active
#     sh3 = wb.create_sheet('update01')
#     for i, row in enumerate(sh1.rows):
#         sh1.cell(i + 1, 4).value = data[i]
#         # row = row.append(data[i])
#         row = [cell.value for cell in row]
#         # print(type(row))
#         # t = data[i]
#         # print(t)
#         row.append(data[i])
#         # print(row)
#         sh3.append(row)
#         print('*')
#     wb.save(tar_file)
#     print('done!')

# 第二种方法：写入多列数据
# from openpyxl import load_workbook
# from datetime import datetime
#
# def create_time():
#     wb= load_workbook(file)
#     sh1 = wb.active
#     lst1 = [cell.value for cell in sh1['D']]
#     lst1 = lst1[1:]
#     curr_year = datetime.now().year
#     str_lst = []
#     str_lst.append('出生日期')
#     age_lst= []
#     age_lst.append('年龄')
#     for i in lst1:
#         print(i)
#         year = int(i[6:10])
#         # print(type(year))
#         month =int(i[10:12])
#         date = int(i[12:14])
#         print(f'{year}年{month}月{date}日')
#         s= ''
#         birt_str= s.join(str(year)+'年'+str(month)+'月'+str(date)+'日')
#         str_lst.append(birt_str)
#         age = curr_year-year
#         age_lst.append(age)
#         # print(f'{age}age')
#     all_data.append(str_lst)
#     all_data.append(age_lst)
#     return all_data
#
# def create_sheet():
#     wb = load_workbook(file,data_only = True)
#     sh1 = wb.active
#     sh3 = wb.create_sheet('update01')
#     for i, row in enumerate(sh1.rows):
#         sh1.cell(i + 1, 5).value = all_data[0][i]
#         sh1.cell(i + 1, 6).value = all_data[1][i]
#         row = [cell.value for cell in row]
#         row[0] = (row[0].date() if type(row[0]) != str else row[0]) # 日期格式的处理
#         row.append(all_data[0][i])
#         row.append(all_data[1][i])
#         sh3.append(row)
#         print('*')
#     wb.save(tar_file)
#     print('done!')
#
#
# all_data = []
# file = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15.xlsx'
# tar_file = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15_output.xlsx'
#
# create_time()
# create_sheet()


# 35 批量发送邮件
#
# import smtplib
# from openpyxl import load_workbook
# from email.mime.text import MIMEText
# from email.header import Header
#
#
# smtp_obj = smtplib.SMTP('smtp.qq.com')
# smtp_obj.login('1551786081@qq.com','yigwusfcctakfijb')
#
# # 编辑内容
# ori_file = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data35.xlsx'
#
# wb = load_workbook(ori_file, data_only = True)
# sh1 = wb.active
#
#
# table_head = '<tr>'
# for i,row in enumerate(sh1.iter_rows()):
#     if i ==0:
#         for col in row:
#             table_head += f'<td>{col.value}</td>'
#         table_head += '</tr>'
#         continue
#     else:
#         table_info = '<tr>'
#         for col in row:
#             table_info += f'<td>{col.value}</td>'
#         table_info += '</tr>'
#
#         name = row[1].value
#         address = row[9].value
#
#         msg_info = f'''
#         <h3>您好，{name}:</h3>
#         <p>请查收2030年5月的工资详情：</p>
#         <table border='1'>
#         {table_head}
#         {table_info}
#         </table>
#         '''
#
#         msg_body = MIMEText(msg_info,'html','utf-8')
#         msg_body['From'] = Header('人力资源部','utf-8')
#         msg_body['To'] = Header('employee')
#         msg_body['Subject'] = Header('都是人才大中华区分公司工资明细','utf-8')
#
#         smtp_obj.sendmail('1551786081@qq.com',address,msg_body.as_string())
#         print(f'工资信息给{name}发送到{address}，成功！')
#


# 32 发送普通邮件

# import smtplib
# from email.mime.text import MIMEText
#
# smtp_obj = smtplib.SMTP('smtp.qq.com')
# smtp_obj.login('1551786081@qq.com','yigwusfcctakfijb')
#
# mail_text = 'This is Email~~~~~'
# msg_body = MIMEText(mail_text,'plain','utf-8')
#
# smtp_obj.sendmail('1551786081@qq.com','405fjl@163.com',msg_body.as_string())
# print('test_mail发送成功！')

# 33 发送html邮件

# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
#
# smtp_obj = smtplib.SMTP('smtp.qq.com')
# smtp_obj.login('1551786081@qq.com','yigwusfcctakfijb')
#
# mail_text = '''
# <h1 style='color:blue'>Title</h1>
# <p>mail text</p>
# <p><a href = 'https://www.baidu.com'>url</a></p>
# <p>Best wishes!</p>
# '''
#
#
# msg_body = MIMEText(mail_text,'html','utf-8')
# msg_body['From'] = Header('测试邮箱','utf-8')
# msg_body['To'] = Header('employee')
# msg_body['Subject'] = Header('测试测试！！！','utf-8')
#
#
# smtp_obj.sendmail('1551786081@qq.com','405fjl@163.com',msg_body.as_string())
# print('test_mail发送成功！')

# 34 发送附件邮件

# 发送单个附件
# import smtplib
# from email.mime.text import MIMEText
# from email.header import Header
# from email.mime.application import MIMEApplication
# from email.mime.multipart import MIMEMultipart
#
# smtp_obj = smtplib.SMTP('smtp.qq.com')
# smtp_obj.login('1551786081@qq.com','yigwusfcctakfijb')
#
# file_path = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15_output.xlsx'
#
# mail_text = '''
# <h1 style='color:blue'>测试附件</h1>
# <p>mail text</p>
# <p><a href = 'https://www.baidu.com'>url</a></p>
# <p>Best wishes!</p>
# '''
#
#
# msg_body = MIMEText(mail_text,'html','utf-8')
# msg_body['From'] = Header('测试邮箱','utf-8')
# msg_body['To'] = Header('employee')
# msg_body['Subject'] = Header('测试测试！！！','utf-8')
#
#
# file = MIMEApplication(open(file_path,'rb').read())
# file.add_header('Content-Disposition','attachment',filename='test_attachment.xlsx')
#
# multi_part = MIMEMultipart()
# multi_part.attach(msg_body)
#
# multi_part.attach(file)
#
# smtp_obj.sendmail('1551786081@qq.com','405fjl@163.com',multi_part.as_string())
# print('test_mail发送成功！')

# 发送多个附件

# 01 引入要用的模块

def send_mail(files):
    import smtplib
    from email.mime.text import MIMEText
    from email.header import Header
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    import os

    # 02 登陆邮箱
    smtp_obj = smtplib.SMTP('smtp.qq.com')
    smtp_obj.login('1551786081@qq.com','yigwusfcctakfijb')


    # 03 编辑邮件内容
    mail_text = '''
    <h1 style='color:blue'>测试附件</h1>
    <p>mail text</p>
    <p><a href = 'https://www.baidu.com'>url</a></p>
    <p>Best wishes!</p>
    '''

    msg_body = MIMEText(mail_text,'html','utf-8')

    # 04 把正文加入multi_part
    multi_part = MIMEMultipart()
    multi_part.attach(msg_body)
    multi_part['From'] = Header('测试邮箱04','utf-8')
    multi_part['To'] = Header('employee04')
    multi_part['Subject'] = Header('测试添加多个附件04','utf-8')

    # 05 for循环加入附件
    # file_path = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15_output.xlsx'


    for name in os.listdir(files):
        ext = os.path.splitext(name)[1]
        if ext == '.xlsx':
            print(name)
            file_path = f'{files}/{name}'
            # print(file_path)

            file = MIMEApplication(open(file_path, 'rb').read())
            file.add_header('Content-Disposition', 'attachment', filename= f'{name}.xlsx')
            multi_part.attach(file)

    print('ready to go!')

    # 06 发送邮件
    smtp_obj.sendmail('1551786081@qq.com','405fjl@163.com',multi_part.as_string())
    print('test_mail发送成功！')

files = '/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09'
send_mail(files)


# if __name__ == '__main__':
    # main()

print('Goodbye')