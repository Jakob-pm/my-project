# date = 03272023
# author = Jakob
print('hello world!')

#8  继续

def set_value():
    from openpyxl.styles import Font, Alignment, colors
    bold_italic_30_font = Font(name = '微软雅黑', size = 30, bold = True, color = colors.BLUE)
    bold_italic_20_font = Font(name='微软雅黑', size=20, italic=True, underline='singleAccounting', bold=True, color='CD1076')
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1['A1'] = 'Hello world!'
    sh1['A1'].font = bold_italic_30_font
    sh1['C3'] = 'Python!'
    sh1['C3'].font = bold_italic_20_font
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')


# new_sheet()
# set_value()

def set_multi_values():
    from openpyxl.styles import Font, Alignment, colors
    # bold_italic_30_font = Font(name='微软雅黑', size=30, bold=True, color=colors.BLUE)
    bold_italic_20_font = Font(name='等线', size=20, italic=True, underline='single', bold=True,
                               color='DAA520')
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    data = ['Python','Java','C++','PHP']
    for i,d in enumerate(data):
        sh1.cell(i+1,1).value = d
        sh1.cell(i+1,1).font = bold_italic_20_font
        print('*')
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')
    print('done')

# set_multi_values()

def set_style2():
    from openpyxl.styles import Font, Alignment, colors
    # bold_italic_30_font = Font(name='微软雅黑', size=30, bold=True, color=colors.BLUE)
    bold_italic_20_font = Font(name='等线', size=20, italic=True, underline='single', bold=True,
                               color='DAA520')
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1.row_dimensions[1].height = 30
    sh1.column_dimensions['B'].width = 50
    data = ['Python', 'Java', 'C++', 'PHP']
    for i, d in enumerate(data):
        sh1.cell(i + 1, 1).value = d
        sh1.cell(i + 1, 1).font = bold_italic_20_font
        sh1.cell(i + 1, 1).alignment = Alignment(horizontal='center',vertical='center')
        print('*')
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')
    print('done')

# set_style2()

def set_merge():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1.merge_cells('A1:C3')
    sh1.merge_cells('D2:E5')
    sh1['A1']='横向合并'
    sh1['D2'] = '多 合并'
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')

# set_merge()

def set_image():
    from openpyxl import Workbook
    from datetime import date
    wb = Workbook()
    sh1 = wb.active
    rows = [
        [date(2023,3,12),40,30,25],
        [date(2023, 3, 13), 40, 30, 25],
        [date(2023, 3, 14), 50, 59, 34],
        [date(2023, 3, 15), 30, 27, 12],
        [date(2023, 3, 16), 20, 56, 57],
        [date(2023, 3, 17), 50, 38, 69]
    ]
    for row in rows:
        sh1.append(row)
    from openpyxl.chart import LineChart,Reference
    c1 = LineChart()
    c1.title = 'LineChart'
    c1.x_axis.title = 'Test_number'
    c1.y_axis.title = 'size'
    data = Reference(sh1,min_col=2,min_row=1,max_col=4,max_row=7)
    c1.add_data(data,titles_from_data = True)
    sh1.add_chart(c1, 'F3')
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08.xlsx')
    print('done')

set_image()

# 09 多个Excel合并为一个excel

# 10 多个Excel合并为一个excel

# 11 生成excel工资条




# if __name__ == '__main__':
    # main()

print('Goodbye')