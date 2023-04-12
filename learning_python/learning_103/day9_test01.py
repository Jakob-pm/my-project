# date = 03/31/2023
# author = Jakob
print('hello world!')

# 15 身份证号提前信息
# 重写一遍，梳理一下逻辑

# 第二种方法：写入多列数据

# 01 引入模块
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment, colors
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side

# 02 获得要插入的数据
# 获得数据的格式为列表套一层列表
'''
[
[],[],[],[]
]
'''

def get_data():
    # 读入数据
    wb= load_workbook(file)
    sh1 = wb.active
    # 拿到对应列的数据，存为列表(-1，去掉表头)
    raw_column = [cell.value for cell in sh1['D']][1:]
    # 当前日期，年份
    curr_year = datetime.now().year
    str_lst = [] # 接收出生年月信息，str
    str_lst.append('出生日期')
    age_lst= [] # 接收年龄信息，int
    age_lst.append('年龄')
    for i in raw_column:
        # print(i)
        year = int(i[6:10])
        # print(type(year))
        month =int(i[10:12])
        date = int(i[12:14])
        # print(f'{year}年{month}月{date}日')
        birt_str= f'{year}年{month}月{date}日' # format的正确用法
        # print(birt_str)
        str_lst.append(birt_str)
        age = curr_year-year
        age_lst.append(age)
        # print(f'{age}age')
    # print(str_lst)
    # print(age_lst)
    all_data.append(str_lst)
    all_data.append(age_lst)
    return all_data

#3 把要插入的数据写入表格
def update_sheet():
    wb = load_workbook(file,data_only = True)
    sh1 = wb.active
    for i, row in enumerate(sh1.rows):
        sh1.cell(i + 1, 5).value = all_data[0][i]
        sh1.cell(i + 1, 6).value = all_data[1][i]
        print('*')
    wb.save(tar_file_01)
    print('testcase01 done!')

def add_sheet():
    wb = load_workbook(file,data_only = True)
    sh1 = wb.active
    sh3 = wb.create_sheet('update01')
    for i, row in enumerate(sh1.rows):
        row = [cell.value for cell in row] # 去掉第一行
        row[0] = (row[0] if i == 0 else row[0].date()) # 日期格式的处理
        row.append(all_data[0][i])
        row.append(all_data[1][i])
        sh3.append(row)
        print('$')

    # 设置单元格格式：
    bold = Font(name='微软雅黑', size=11,  bold=True)
    fill = PatternFill('solid', fgColor='C0C0C0')
    borders = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'))
    sh3.row_dimensions[1].height = 19
    sh3.column_dimensions['A'].width = 15
    sh3.column_dimensions['D'].width = 20
    sh3.column_dimensions['E'].width = 15
    sh3.column_dimensions['F'].width = 5
    for cell in sh3[1]:
        cell.font = bold
        cell.fill = fill

    for i, row in enumerate(sh3.rows):
        for k, cell in enumerate(row):
            # print(cell)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borders

    del wb['Sheet']
    wb.save(tar_file_02)
    print('testcase02 done!')


all_data = []
file = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15.xlsx'
tar_file_01 = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15_testcase01.xlsx'
tar_file_02 = '/Users/jakob_pc/Desktop/Python_datas/103_day9/data15_testcase02.xlsx'


# create_time()
get_data()
# create_sheet()
# update_sheet()
add_sheet()
print('both done!')




print('Goodbye')

