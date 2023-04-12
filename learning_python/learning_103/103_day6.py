# date = 03282023
# author = Jakob
print('hello world!')
#8  继续

'''def set_image2():
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    # from openpyxl.chart.series import DataPoint

    data = [
        ['名称','数值'],
        ['苹果', 50],
        ['草莓', 30],
        ['椰子', 10],
        ['荔枝', 40]
    ]
    wb = Workbook()
    sh1 = wb.active

    for row in data:
        sh1.append(row)

    pie = PieChart()
    labels = Reference(sh1,min_col=1,min_row=2,max_row=5)
    data = Reference(sh1,min_col=2,min_row=1,max_row=5)
    pie.add_data(data, titles_from_data = True)
    pie.set_categories(labels)
    pie.title = 'Pie sold by category'

    sh1.add_chart(pie, 'D1')
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08_01.xlsx')
    print('done')

# set_image2()

def set_image3():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Series, Reference
    wb = Workbook()
    sh1 = wb.active

    rows = [
        ('Number','Batch1','Batch2'),
        (2, 10, 30),
        (3, 40, 60),
        (4, 50, 70),
        (5, 80, 20),
        (6, 30, 40),
        (7, 50, 90)
    ]

    for row in rows:
        sh1.append(row)
        chart1 = BarChart()
        chart1.type = 'col'
        chart1.style = 10 #不理解，颜色
        chart1.title = 'Bar Chart'
        chart1.y_axis.title = 'Test Number'
        chart1.x_axis.title = 'Sample length (mm)'

        data = Reference(sh1, min_col=2, min_row=1, max_row=7,max_col=3)
        cats = Reference(sh1, min_col=1, min_row=2, max_row=7)
        chart1.add_data(data, titles_from_data = True)
        chart1.set_categories(cats)
        chart1.shape = 4
    sh1.add_chart(chart1,'A10')
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_test_08_02.xlsx')
    print('done')

set_image3()'''


# 09 多个Excel合并为一个excel

# 原始数据路径：/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09

'''from openpyxl import load_workbook, Workbook
import os

def merge_data():
    wb = Workbook()
    sh1 = wb.active
    all_data = []
    for name in os.listdir('/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09'):
        ext = os.path.splitext(name)[1]
        if ext == '.xlsx':
            print(name)
            file_path =f'/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09/{name}'
            temp_wb = load_workbook(file_path)
            temp_sh1 = temp_wb.active
            for r in range(1, temp_sh1.max_row+1):
                row_value = []
                for c in range(1,temp_sh1.max_column+1):
                    value = temp_sh1.cell(r,c).value
                    row_value.append(value)
                if row_value not in all_data:
                    all_data.append(row_value)
    for data in all_data:
        sh1.append(data)
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09/merge_data.xlsx')
    print('done!')

merge_data()
'''

# 10 多个Excel合并为一个excel的多个sheet

'''from openpyxl import load_workbook, Workbook
import os

def merge_data2():
    wb = Workbook()
    for name in os.listdir('/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09'):
        ext = os.path.splitext(name)[1]
        if ext == '.xlsx':
            print(name)
            file_path =f'/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09/{name}'
            temp_wb = load_workbook(file_path)
            temp_sh1 = temp_wb.active
            name = name.replace(ext,'')
            sh1 = wb.create_sheet(name)
            for r in range(1, temp_sh1.max_row+1):
                row = []
                for c in range(1,temp_sh1.max_column+1):
                    value = temp_sh1.cell(r,c).value
                    row.append(value)
                sh1.append(row)
    del wb['Sheet']
    wb.save('/Users/jakob_pc/Desktop/Python_datas/103_day6/data_09/merge_data.xlsx')
    print('done!')

merge_data2()
'''


# 11 生成excel工资条





# if __name__ == '__main__':
    # main()

print('Goodbye')